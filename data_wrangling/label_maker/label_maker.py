from openpyxl import load_workbook

student_data = "../labels_test.xlsx"
blank_labels = "../labels_blank.xlsx"
outfile = "../labels_out.xlsx"

def excel_to_list(path):
    labels = []
    wb = load_workbook(filename=path, read_only=True)
    ws = wb.active
    for row in range (2, 163):
        temp_list = [ws["A"+str(row)].value, ws["B"+str(row)].value, ws["C"+str(row)].value]
        labels.append(temp_list)
    return labels

def list_to_labels(list, path, out):
    BT, BM, BB = 2, 3, 4
    DT, DM, DB = 2, 3, 4
    FT, FM, FB = 2, 3, 4
    HT, HM, HB = 2, 3, 4
    JT, JM, JB = 2, 3, 4
    LT, LM, LB = 2, 3, 4
    NT, NM, NB = 2, 3, 4
    wb = load_workbook(filename=path)
    ws = wb.active
    for i in list:
        if BT < 91 and BM < 92 and BB <93:
            ws["B"+str(BT)] = i[0]
            BT = BT+4
            ws["B"+str(BM)] = i[1]
            BM = BM+4
            ws["B"+str(BB)] = i[2]
            BB = BB+4
        elif DT < 91 and DM < 92 and DB <93:
            ws["D"+str(DT)] = i[0]
            DT = DT+4
            ws["D"+str(DM)] = i[1]
            DM = DM+4
            ws["D"+str(DB)] = i[2]
            DB = DB+4
        elif FT < 91 and FM < 92 and FB <93:
            ws["F"+str(FT)] = i[0]
            FT = FT+4
            ws["F"+str(FM)] = i[1]
            FM = FM+4
            ws["F"+str(FB)] = i[2]
            FB = FB+4
        elif HT < 91 and HM < 92 and HB <93:
            ws["H"+str(HT)] = i[0]
            HT = HT+4
            ws["H"+str(HM)] = i[1]
            HM = HM+4
            ws["H"+str(HB)] = i[2]
            HB = HB+4
        elif JT < 91 and JM < 92 and JB <93:
            ws["J"+str(JT)] = i[0]
            JT = JT+4
            ws["J"+str(JM)] = i[1]
            JM = JM+4
            ws["J"+str(JB)] = i[2]
            JB = JB+4
        elif LT < 91 and LM < 92 and LB <93:
            ws["L"+str(LT)] = i[0]
            LT = LT+4
            ws["L"+str(LM)] = i[1]
            LM = LM+4
            ws["L"+str(LB)] = i[2]
            LB = LB+4
        elif NT < 91 and NM < 92 and NB <93:
            ws["N"+str(NT)] = i[0]
            NT = NT+4
            ws["N"+str(NM)] = i[1]
            NM = NM+4
            ws["N"+str(NB)] = i[2]
            NB = NB+4
    wb.save(filename = out)

label_list = excel_to_list(student_data)
list_to_labels(label_list, blank_labels, outfile)